import os
import sys
# import pickle
from tkinter import filedialog
from tkinter import Tk
sys.path.insert(0, './lib') #to open user-defined functions
# from m_open_extension import read_pickle
from argparse import ArgumentParser
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from m_open_extension import *
from m_fft import *

from openpyxl import load_workbook

import matplotlib
import cmath
from scipy import signal
from Burst_Detection import burst_detector



Inputs = ['mode', 'fs', 'channel']
InputsOpt_Defaults = {'value':1, 'us_before':1000, 'us_after':5000}

def main(argv):
	config = read_parser(argv, Inputs, InputsOpt_Defaults)
	
	if config['mode'] == 'simple_plot':
		root = Tk()
		root.withdraw()
		root.update()
		filepath = filedialog.askopenfilename()			
		root.destroy()
		data = load_signal(filepath, channel=config['channel'])
		t = [i/config['fs'] for i in range(len(data))]
		# plt.xlabel('Time s')
		# plt.ylabel('Amplitude V')
		# plt.title('Z2, S3, Rep. 1, V2')
		plt.figure(1)
		plt.plot(t, data)
		
		plt.figure(2)
		magX, f, df = mag_fft(data, config['fs'])
		plt.plot(f, magX, 'r')
		
		
		plt.show()			
		
	elif config['mode'] == 'two_correlation':
		root = Tk()
		root.withdraw()
		root.update()
		Filepaths = filedialog.askopenfilenames()			
		root.destroy()
		Data = [load_signal(filepath, channel=config['channel']) for filepath in Filepaths]
		
		
		# Correlation = np.correlate(Data[0], Data[1], mode='same')
		# data = load_signal(filepath, channel=config['channel'])

		
		Data[0] = Data[0][int(np.argmax(Data[0]) - 1e5):int(np.argmax(Data[0]) + 1e5)]
		Data[1] = Data[1][int(np.argmax(Data[1]) - 1e5):int(np.argmax(Data[1]) + 1e5)]
		# for k in range(2):
			# f_psd, Data[k] = signal.periodogram(Data[k], config['fs'], return_onesided=True, scaling='density')
		
		Correlation = np.correlate(Data[0]/(np.sum(Data[0]**2))**0.5, Data[1]/(np.sum(Data[1]**2))**0.5, mode='same')
		# t = [i/config['fs'] for i in range(len(Data[0]))]
		
		# fig, ax = plt.subplots(nrows=2, ncols=1, sharex=True, sharey=True)
		# ax[0].ticklabel_format(axis='y', style='sci', scilimits=(-2, 2))
		# ax[1].ticklabel_format(axis='y', style='sci', scilimits=(-2, 2))
		
		# # plt.figure(0)
		# # ax[0].plot(np.array(f_psd)/1000., Data[0], 'g')	
		# ax[0].plot(t, Data[0])			
		# ax[0].set_ylabel('PSD $\mathregular{V^{2}}$/Hz')
		# # ax[0].set_ylabel('Amplitude V')
		# ax[0].set_title('Burst on S3, Rep. 1, V3')
		
		# # plt.figure(1)
		# ax[1].plot(np.array(f_psd)/1000., Data[1], 'g')
		# # ax[1].plot(t, Data[1])	
		# # ax[1].set_xlabel('Rel. Time s')
		# ax[1].set_xlabel('Frequency kHz')
		# ax[1].set_ylabel('PSD $\mathregular{V^{2}}$/Hz')
		# # ax[1].set_ylabel('Amplitude V')
		# ax[1].set_title('Burst on S3, Rep. 3, V3')
		
		
		
		
		# plt.figure(1)
		# plt.plot(Data[1])
		# plt.figure(2)
		# plt.plot(f_psd, Correlation)
		print(np.max(Correlation))

	
	elif config['mode'] == 'reference_correlation':
		print('Select Reference Burst')
		root = Tk()
		root.withdraw()
		root.update()
		filepath_Ref = filedialog.askopenfilename()			
		root.destroy()
		data_ref = load_signal(filepath_Ref, channel=config['channel'])
	
		print('Select Burst to Correlate')
		root = Tk()
		root.withdraw()
		root.update()
		Filepaths = filedialog.askopenfilenames()			
		root.destroy()
		Data = [load_signal(filepath, channel=config['channel']) for filepath in Filepaths]		

		
		
		data_ref = data_ref[int(np.argmax(data_ref) - config['us_before']*config['fs']/1.e6):int(np.argmax(data_ref) + config['us_after']*config['fs']/1.e6)]
		
		
		# Data[0] = Data[0][int(np.argmax(Data[0]) - config['us_before']*config['fs']/1.e6):int(np.argmax(Data[0]) + config['us_after']*config['fs']/1.e6)]
		# Data[1] = Data[1][int(np.argmax(Data[1]) - config['us_before']*config['fs']/1.e6):int(np.argmax(Data[1]) + config['us_after']*config['fs']/1.e6)]
		t = [i/config['fs'] for i in range(len(data_ref))]
		# plt.plot(data_ref)
		# plt.show()
		
		# for k in range(2):
			# f_psd, Data[k] = signal.periodogram(Data[k], config['fs'], return_onesided=True, scaling='density')
		
		# Correlation = np.correlate(Data[0]/(np.sum(Data[0]**2))**0.5, Data[1]/(np.sum(Data[1]**2))**0.5, mode='same')
		#
		for data in Data:
			data = data[int(np.argmax(data) - config['us_before']*config['fs']/1.e6):int(np.argmax(data) + config['us_after']*config['fs']/1.e6)]
			print(max_norm_correlation(data_ref, data))
			print(argmax_norm_correlation(data_ref, data))
			print(len(data))
			plt.plot(data)
			plt.plot(data_ref, 'r')
			plt.show()

	
	elif config['mode'] == 'plot_interp':
		x = np.array([215, 880, 1670])
		y = np.array([307, 57, 27])
		# y = np.log(y)
		from scipy.interpolate import interp1d
		# f = interp1d(x, y)
		# f2 = interp1d(x, y, kind='quadratic')
		# xnew = np.linspace(215, 1670, num=40, endpoint=True)
		# plt.plot(x, y, 'o', xnew, f(xnew), '-', xnew, f2(xnew), '--')
		# plt.legend(['data', 'linear', 'cubic'], loc='best')
		# plt.xlabel('Distance (mm)')
		# plt.ylabel('Amplitude over Noise (mV)')
		# plt.xlim((100, 2000))
		# plt.show()
		print(y)
		z2 = np.polyfit(x, np.log(y), 1)
		z = np.polyfit(x, y, 1)	
		p = np.poly1d(z)
		p2 = np.poly1d(z2)
		xp = np.linspace(50, 2500, num=40, endpoint=True)
		plt.plot(x, y, 'o', xp, p(xp), '-', xp, np.exp(1.02*p2(xp))-10, '-')
		plt.legend(['data', 'linear', 'exp'], loc='best')
		# plt.plot(x, y, '.', xp, np.exp(p(xp)), '-')
		plt.xlabel('Distance (mm)')
		plt.ylabel('Amplitude over Noise (mV)')
		plt.xlim((50, 2500))
		plt.ylim((0, 350))

		
		plt.show()

	elif config['mode'] == 'Xcorr_analysis':
		mydict = {}
		for count in range(3):
			count = count + 1
			maxxcorr_values = []
			print('++++++++Select 5 repetitions for the Test N°', count)
			root = Tk()
			root.withdraw()
			root.update()
			Filepaths = filedialog.askopenfilenames()			
			root.destroy()
			Data = [load_signal(filepath, channel=config['channel']) for filepath in Filepaths]

			for k in range(5):
				Data[k] = Data[k][int(np.argmax(Data[k]) - 1e5):int(np.argmax(Data[k]) + 1e5)]			
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[1]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[2]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[2]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[2], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[2], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[3], Data[4]))
			
			mydict['V' + str(count)] = maxxcorr_values
		
		row_names = ['Max_Xcorr_12', 'Max_Xcorr_13', 'Max_Xcorr_14', 'Max_Xcorr_15', 'Max_Xcorr_23', 'Max_Xcorr_24', 'Max_Xcorr_25', 'Max_Xcorr_34', 'Max_Xcorr_35', 'Max_Xcorr_45']
		DataFr = pd.DataFrame(data=mydict, index=row_names)
		# mydict = {'12', '13', '14', '15', '23', '24', '25', '34', '35', '45'}
		# book = load_workbook('Results_Xcorr_Test.xlsx')		
		writer = pd.ExcelWriter('to_use.xlsx')
		# writer.book = book
		# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		
		DataFr.to_excel(writer, sheet_name='Sheet')	
	
	elif config['mode'] == 'bubble_polar_plot':

		RMS = [1.57e-2, 5.45e-3, 6.22e-3]
		
		
		zahn = 2
		
		rms_sum = np.sum(np.array(RMS))
		RMS_rel = RMS / rms_sum


		
		angles = np.array([78, 318, 198])*2*3.14/360
		
		
		
		radii = RMS_rel
		
		
		vec0_rect = cmath.rect(RMS_rel[0], angles[0])
		vec1_rect = cmath.rect(RMS_rel[1], angles[1])
		vec2_rect = cmath.rect(RMS_rel[2], angles[2])
		vecS_rect = vec0_rect + vec1_rect + vec2_rect
		vecS_pol = cmath.polar(vecS_rect)	

		
		result_angle = vecS_pol[1]
		result_radii = vecS_pol[0]		
		
		
			
		fig, ax = plt.subplots(subplot_kw=dict(polar=True),)
		kw = dict(arrowstyle="->", color='k', lw=2)
		
		
		# if arrowprops:
			# kw.update(arrowprops)
		# [ax.annotate("", xy=(angle, radius), xytext=(0, 0),
					 # arrowprops=kw) for
		 # angle, radius in zip(angles, radii)]		 
		 
		 
		kw = dict(arrowstyle="->", color='r', lw=2)
		
		
		# xy = (78*2*3.14/360, 0.5)
		# [ax.annotate("", xy=xy, xytext=(0, 0),
					 # arrowprops=kw) for
		 # angle, radius in zip(angles, radii)]
		 
		 
		# ax.set_ylim(0, np.max(radii))	


		
		ax.set_facecolor('lavender')
		ax.set_rmax(1.0)
		# ax.text(x=angles[0], y=1.05, s='S1', fontsize=20)
		# ax.text(x=angles[1], y=1.1, s='S2', fontsize=20)
		ax.text(x=angles[2], y=1.24, s='S4', fontsize=20)
		ax.set_ylim(0, 1)
		
		
		
		# if zahn == 1:
			# plt.scatter((90+12)*2*np.pi/360, 1, s=200, c='limegreen')
		# elif zahn == 2:
			# plt.scatter(90*2*np.pi/360, 1, s=200, c='g')
		# elif zahn == 3:		
			# plt.scatter((90-12)*2*np.pi/360, 1, s=200, c='g')
			
		# plt.scatter((355.5)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((23)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((50.5)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((105.5)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((133)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((160.5)*2*np.pi/360, 1, s=200, c='r')
		# plt.scatter((55.)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((40.)*2*np.pi/360, 1, s=200, c='g')
		# plt.scatter((90.)*2*np.pi/360, 1, s=200, c='r')
		
		# ax.text(x=(355.5)*2*np.pi/360, y=1.05, s='A', fontsize=20)
		# ax.text(x=(23)*2*np.pi/360, y=1.07, s='B', fontsize=20)
		# ax.text(x=(50.5)*2*np.pi/360, y=1.07, s='C', fontsize=20)
		# ax.text(x=(105.5)*2*np.pi/360, y=1.07, s='D', fontsize=20)
		# ax.text(x=(133)*2*np.pi/360, y=1.1, s='E', fontsize=20)
		# ax.text(x=(160.5)*2*np.pi/360, y=1.1, s='F', fontsize=20)
		# ax.text(x=(55.)*2*np.pi/360, y=1.1, s='Tooth 1', fontsize=20)
		# ax.text(x=(40.)*2*np.pi/360, y=1.1, s='Tooth 2', fontsize=20)
		# ax.text(x=(90.)*2*np.pi/360, y=1.1, s='O', fontsize=20)
		
		ax.set_xticks([])
		ax.set_yticks([])
		
		
		# plt.scatter(angles[0], 1, s=200, c='b')
		# plt.scatter(angles[1], 1, s=200, c='b')
		plt.scatter(angles[2], 1, s=200, c='b')
		# plt.scatter(result_angle, 1, s=200, c='r')
		
		plt.show()
	
	elif config['mode'] == 'XcorrSpectral_analysis':	
		
		
		mydict = {}
		print('Cross Spectral Density')
		for count in range(3):
			count = count + 1
			maxxcorr_values = []
			print('++++++++Select 5 repetitions for the Test N°', count)
			root = Tk()
			root.withdraw()
			root.update()
			Filepaths = filedialog.askopenfilenames()			
			root.destroy()
			Data = [load_signal(filepath, channel=config['channel']) for filepath in Filepaths]

			for k in range(5):				
				Data[k] = Data[k][int(np.argmax(Data[k]) - 1e5):int(np.argmax(Data[k]) + 1e5)]
				f_psd, Data[k] = signal.periodogram(Data[k], config['fs'], return_onesided=True, scaling='density')
				
			# plt.plot(f_psd, Data[k])
			# plt.show()
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[1]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[2]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[0], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[2]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[1], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[2], Data[3]))
			maxxcorr_values.append(max_norm_correlation(Data[2], Data[4]))
			maxxcorr_values.append(max_norm_correlation(Data[3], Data[4]))
			
			mydict['V' + str(count)] = maxxcorr_values
		
		row_names = ['Max_Xcorr_12', 'Max_Xcorr_13', 'Max_Xcorr_14', 'Max_Xcorr_15', 'Max_Xcorr_23', 'Max_Xcorr_24', 'Max_Xcorr_25', 'Max_Xcorr_34', 'Max_Xcorr_35', 'Max_Xcorr_45']
		DataFr = pd.DataFrame(data=mydict, index=row_names)
		# mydict = {'12', '13', '14', '15', '23', '24', '25', '34', '35', '45'}
		# book = load_workbook('Results_Xcorr_Test.xlsx')		
		writer = pd.ExcelWriter('to_use.xlsx')
		# writer.book = book
		# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		
		DataFr.to_excel(writer, sheet_name='CSD')	
		
		
	elif config['mode'] == 'three_plot':
		root = Tk()
		root.withdraw()
		root.update()
		filepath = filedialog.askopenfilename()			
		root.destroy()
		
		if config['channel'] == 'all':
			Channels = ['0', '1', '2']
		Data = [load_signal(filepath, channel=channel) for channel in Channels]
		
		# Data[0] = Data[0][int(np.argmax(Data[0]) - 1e3):int(np.argmax(Data[0]) + 1e5)]
		# Data[1] = Data[1][int(np.argmax(Data[1]) - 1e3):int(np.argmax(Data[1]) + 1e5)]
		# Data[2] = Data[2][int(np.argmax(Data[2]) - 1e3):int(np.argmax(Data[2]) + 1e5)]

		
		# Correlation = np.correlate(Data[0]/(np.sum(Data[0]**2))**0.5, Data[1]/(np.sum(Data[1]**2))**0.5, mode='same')
		
		
		# fig, ax = plt.subplots(nrows=3, ncols=1, sharex=True, sharey=True)
		# ax[0].ticklabel_format(axis='y', style='sci', scilimits=(-2, 2))
		# ax[1].ticklabel_format(axis='y', style='sci', scilimits=(-2, 2))
		# ax[2].ticklabel_format(axis='y', style='sci', scilimits=(-2, 2))
		
		# ax[0].plot(t, Data[0])			
		# ax[0].set_ylabel('Amplitude V')
		# ax[0].set_title('Sensor 1', fontsize=10)
		

		# ax[1].plot(t, Data[1])			
		# ax[1].set_ylabel('Amplitude V')
		# ax[1].set_title('Sensor 2', fontsize=10)
		
		# ax[2].plot(t, Data[2])			
		# ax[2].set_ylabel('Amplitude V')
		# ax[2].set_title('Sensor 3', fontsize=10)
		# ax[2].set_xlabel('Time s')
		
		
		
		# plt.show()
		t = [i/config['fs'] for i in range(len(Data[0]))]
		
	elif config['mode'] == 'localization':

		mydict = {}
		print('Localization Analysis')
		for count in range(3):
			count = count + 1
			angle_values = []
			print('++++++++Select 5 repetitions for the Teeth N°', count)
			root = Tk()
			root.withdraw()
			root.update()
			Filepaths = filedialog.askopenfilenames()			
			root.destroy()
			if config['channel'] == 'all':
				Channels = ['0', '1', '2']
			

			# for k in range(5):				
				# Data[k] = Data[k][int(np.argmax(Data[k]) - 1e3):int(np.argmax(Data[k]) + 1e5)]
				# f_psd, Data[k] = signal.periodogram(Data[k], config['fs'], return_onesided=True, scaling='density')
				
			# plt.plot(f_psd, Data[k])
			# plt.show()
			for filepath in Filepaths:
				Data = [load_signal(filepath, channel=channel) for channel in Channels]
				angle_values.append(calculate_angle(Data[0], Data[1], Data[2], config))


			
			mydict['Teeth ' + str(count)] = angle_values
		
		row_names = ['Angle_Rep_1', 'Angle_Rep_2', 'Angle_Rep_3', 'Angle_Rep_4', 'Angle_Rep_5']
		DataFr = pd.DataFrame(data=mydict, index=row_names)
		# mydict = {'12', '13', '14', '15', '23', '24', '25', '34', '35', '45'}
		# book = load_workbook('Results_Xcorr_Test.xlsx')		
		writer = pd.ExcelWriter('to_use.xlsx')
		# writer.book = book
		# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		
		DataFr.to_excel(writer, sheet_name='Angles')	
	
		
	else:
		print('unknown mode')
		sys.exit()

		
		

		
		
	return


def compass(u, v, arrowprops=None):
    """
    Compass draws a graph that displays the vectors with
    components `u` and `v` as arrows from the origin.

    Examples
    --------
    >>> import numpy as np
    >>> u = [+0, +0.5, -0.50, -0.90]
    >>> v = [+1, +0.5, -0.45, +0.85]
    >>> compass(u, v)
    """

    # angles, radii = cart2pol(u, v)

    fig, ax = plt.subplots(subplot_kw=dict(polar=True))

    kw = dict(arrowstyle="->", color='k')
    if arrowprops:
        kw.update(arrowprops)
    [ax.annotate("", xy=(angle, radius), xytext=(0, 0),
                 arrowprops=kw) for
     angle, radius in zip(angles, radii)]

    ax.set_ylim(0, np.max(radii))

    return fig, ax

def cart2pol(x, y):
    """Convert from Cartesian to polar coordinates.

    Example
    -------
    >>> theta, radius = pol2cart(x, y)
    """
    radius = np.hypot(x, y)
    theta = np.arctan2(y, x)
    return theta, radius

def read_parser(argv, Inputs, InputsOpt_Defaults):
	Inputs_opt = [key for key in InputsOpt_Defaults]
	Defaults = [InputsOpt_Defaults[key] for key in InputsOpt_Defaults]
	parser = ArgumentParser()
	for element in (Inputs + Inputs_opt):
		print(element)
		if element == 'no_element':
			parser.add_argument('--' + element, nargs='+')
		else:
			parser.add_argument('--' + element, nargs='?')
	
	args = parser.parse_args()
	config = {}
	for element in Inputs:
		if getattr(args, element) != None:
			config[element] = getattr(args, element)
		else:
			print('Required:', element)
			sys.exit()

	for element, value in zip(Inputs_opt, Defaults):
		if getattr(args, element) != None:
			config[element] = getattr(args, element)
		else:
			print('Default ' + element + ' = ', value)
			config[element] = value
	
	#Type conversion to float
	config['value'] = float(config['value'])
	config['fs'] = float(config['fs'])
	# config['fscore_min'] = float(config['fscore_min'])
	#Type conversion to int	
	# Variable conversion
	return config


def calculate_angle(signal1, signal2, signal3, config):
		
	config['clf_check'] = 'OFF'
	config['method'] = 'EDG'
	# config['rms_change'] = 8.2
	config['rms_change'] = 4.2
	config['feat_norm'] = 'standard'
	config['features'] = 'DataSorted'
	config['data_norm'] = 'per_rms'
	config['denois'] = 'OFF'
	config['processing'] = 'butter_demod'
	config['diff'] = 1
	config['window_time'] = 0.001
	config['overlap'] = 0
	config['window_delay'] = 0
	config['EMD'] = 'OFF'
	config['NN_model'] = 'C:\\Felix\\Data\\CNs_Getriebe\\Paper_Bursts\\NN_Models\\clf_20171018_180003.pkl'
	config['demod_filter'] = ['lowpass', 2000., 3]
	config['demod_prefilter'] = ['highpass', 70.e3, 3]
	# config['demod_prefilter'] = 'OFF'
	config['demod_rect'] = None
	config['demod_dc'] = None
	config['class2'] = 0
	config['classes'] = '3n_2isclass'
	config['thr_mode'] = 'fixed_value'
	config['thr_value'] = 0.0004
	
	
	x1, t_burst_corr1, amp_burst_corr1, Results, clf_1 = burst_detector(signal1, config, count=None)
	# x1 = x1[10000:]
	t = [i/config['fs'] for i in range(len(x1))]
	# plt.plot(x1)
	# plt.show()
	# t1 = t_burst_corr1[0]
	t1 = np.argmax(x1[5000:])/config['fs']
	
	x1, t_burst_corr1, amp_burst_corr1, Results, clf_1 = burst_detector(signal2, config, count=None)
	# x1 = x1[10000:]
	# plt.plot(x1)
	# plt.show()
	# t2 = t_burst_corr1[0]
	t2 = np.argmax(x1[5000:])/config['fs']
	
	x1, t_burst_corr1, amp_burst_corr1, Results, clf_1 = burst_detector(signal3, config, count=None)
	# x1 = x1[10000:]
	# plt.plot(x1)
	# plt.show()
	# t3 = t_burst_corr1[0]
	t3 = np.argmax(x1[5000:])/config['fs']
	
	

	

	# t1 = np.argmax(Data[0])/config['fs']
	# t2 = np.argmax(Data[1])/config['fs']
	# t3 = np.argmax(Data[2])/config['fs']
	

	
	t12 = t2 - t1
	t13 = t3 - t1
	theta1 = 57
	theta2 = 57 + 240
	theta3 = 57 + 120
	theta13 = theta3 - theta1
	theta12 = 120
	theta01 = (theta13*t12/theta12 - t13)*theta12/(2*t12)
	theta0 = theta01 + theta1
	# print('Time of Max in Sensor 1: ', t1)
	# print('Time of Max in Sensor 2: ', t2)
	# print('Time of Max in Sensor 3: ', t3)
	# print('dt 1-2: ', t12)
	# print('dt 1-3: ', t13)
	# print('Theta 1: ', theta1)
	# print('Theta 2: ', theta2)
	# print('Theta 3: ', theta3)
	# print('Theta 1-3: ', theta13)
	# print('Theta 1-2: ', theta12)
	# print('Theta 0-1: ', theta01)
	# print('Theta 0: ', theta0)
	return theta0


def five_burst_correlation(Data):
	correlation
	return

def max_norm_correlation(signal1, signal2):
	correlation = np.correlate(signal1/(np.sum(signal1**2))**0.5, signal2/(np.sum(signal2**2))**0.5, mode='same')
	return np.max(correlation)

def argmax_norm_correlation(signal1, signal2):
	correlation = np.correlate(signal1/(np.sum(signal1**2))**0.5, signal2/(np.sum(signal2**2))**0.5, mode='same')
	return np.argmax(correlation)
	
if __name__ == '__main__':
	main(sys.argv)
